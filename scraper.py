import sys

import requests
import settings

from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook, Workbook
from datetime import datetime


class Page:
    def __init__(self):
        self.url = 'https://www.bankinfosecurity.com/resources/p-%s'
        self.p = 1
        self._set_html()
        self._set_links()

    def _set_html(self):
        self.html = requests.get(self.url % self.p).text

    def _set_links(self):
        soup = bs(self.html, 'html5lib')
        resources = soup.find_all('h2', {'class': 'title top-none'})
        links = []
        for resource in resources:
            for child in resource.children:
                links.append(child.attrs['href'])
        self.links = links

    def next_page(self):
        self.p += 1
        self._set_html()
        self._set_links()

    def get_links(self):
        return self.links


class Resource:
    def __init__(self, link):
        self.asset_name = None
        self.company_name = None
        self.date = None
        self.type_of_asset = None

        self.link = link
        self._set_html()
        self._parse_data_from_html()

    def _set_html(self):
        self.html = requests.get(self.link).text

    def _parse_data_from_html(self):
        soup = bs(self.html, 'html5lib')
        self.asset_name = soup.find('a', {'class': 'article-title__link'}).contents[0]
        article_byline = soup.find('span', {'class': 'article-byline'}).children
        self.company_name = next(article_byline).strip(' \n•')
        self.date = next(article_byline).contents[0]
        self.type_of_asset = soup.find('a', {'id': 'dld_btn'}).contents[0].strip('\n Download')

    def get_data(self):
        return {'Asset': self.asset_name,
                'Asset link': self.link,
                'Company': self.company_name,
                'Type': self.type_of_asset,
                'Date Posted': self.date}


def write_data_to_excel(data):
    created = False

    if not settings.NEW_XLSX:
        try:
            wb = load_workbook(filename=settings.XLSX_FILENAME)
        except FileNotFoundError:
            wb = Workbook()
            created = True
    else:
        wb = Workbook()
        created = True

    ws = wb.active

    if created:
        ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'], ws['F1'], ws['G1'] = 'Platform', 'Asset', 'Asset_link', 'Company', 'Type', 'Date Posted', 'GEO'

    length = len(ws['A'])
    for i in range(length + 1, len(data) + length + 1):
        ws[f'A{i}'], ws[f'B{i}'], ws[f'C{i}'] = 'BankInfoSecurity from ISMG', data[i - length - 1]['Asset'], data[i - length - 1]['Asset link']
        ws[f'D{i}'], ws[f'E{i}'], ws[f'F{i}'] = data[i - length - 1]['Company'], data[i - length - 1]['Type'], data[i - length - 1]['Date Posted']
        ws[f'G{i}'] = 'USA'

    wb.save(settings.XLSX_FILENAME)


def get_stop_link():
    try:
        with open('stop_link.txt', 'r') as f:
            stop_link = f.readline().strip()
    except FileNotFoundError:
        stop_link = ''
    return stop_link


def update_stop_link(link):
    with open('stop_link.txt', 'w') as f:
        f.write(link)


def main():
    try:
        date_input = sys.argv[1]
        try:
            date_obj = datetime.strptime(date_input, '%m/%d/%Y')
            date_limit = date_obj.date()
        except ValueError:
            exit('Please, enter date in format "mm/dd/yyyy"')
        except Exception as e:
            print('An unexpected error occurred while processing the date limit. Error description:', e)
            exit('Exited.')
    except IndexError:
        date_limit = False

    data = []
    stop_link = get_stop_link()
    main_page = Page()

    count = 0
    print(f'loaded {count} assets', end='')
    while main_page.get_links():
        for link in main_page.get_links():
            resource = Resource(link)
            if link == stop_link and not date_limit or \
                    date_limit and date_limit > datetime.strptime(resource.get_data()['Date Posted'], '%B %d, %Y').date():
                if data:
                    write_data_to_excel(data)
                    update_stop_link(data[0]['Asset link'])
                    print(f'\nSuccessfully saved under "{settings.XLSX_FILENAME}"')
                else:
                    print('\nNo new articles were scraped!')
                return
            data.append(resource.get_data())
            count += 1
            print('\r', end='')
            print(f'loaded {count} assets', end='')

        main_page.next_page()

    if data:
        write_data_to_excel(data)
        update_stop_link(data[0]['Asset link'])
        print(f'\nSuccessfully saved under "{settings.XLSX_FILENAME}"')
    else:
        print('\nNo new articles were scraped!')


if __name__ == '__main__':
    main()
